from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from validator.layout import Layout
from validator.patterns import load_patterns
from validator.validator import (
    grouped_errors_payload,
    validate_workbook,
    write_error_cell_highlights,
)


class LayoutLoadTests(unittest.TestCase):
    def test_load_keeps_column_type_metadata(self) -> None:
        layout_payload = {
            "initialLine": 2,
            "columns": {
                "A": {"description": "Nome", "type": "text", "required": True, "patterns": "ER1"},
                "B": {"description": "UF", "required": False, "domain": "D1"},
            },
        }
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "layout.json"
            path.write_text(json.dumps(layout_payload), encoding="utf-8")
            layout = Layout.load(path)
        self.assertEqual(layout.initial_line, 2)
        self.assertIn("A", layout.required_columns)
        self.assertEqual(layout.column_types.get("A"), "text")
        self.assertEqual(layout.rules["A"], ["ER1"])
        self.assertEqual(layout.domain_rules["B"], "D1")


class PatternsLoadTests(unittest.TestCase):
    def test_load_patterns_parses_regexes_and_domains(self) -> None:
        payload = {
            "patterns": {"ER1": r"^\d{3}$"},
            "domains": {"D1": ["SP", "RJ"]},
        }
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "patterns.json"
            path.write_text(json.dumps(payload), encoding="utf-8")
            compiled = load_patterns(path)
        self.assertIn("ER1", compiled.regexes)
        self.assertIn("D1", compiled.domains)
        self.assertIn("SP", compiled.domains["D1"])


class WorkbookValidationTests(unittest.TestCase):
    def test_validate_workbook_reports_required_and_domain_errors(self) -> None:
        layout_payload = {
            "initialLine": 2,
            "columns": {
                "A": {"description": "CPF", "type": "text", "required": True, "patterns": "ER1"},
                "B": {"description": "UF", "required": True, "domain": "D1"},
            },
        }
        patterns_payload = {
            "patterns": {"ER1": r"^\d{11}$"},
            "domains": {"D1": ["SP", "RJ"]},
        }

        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            workbook_path = base / "sample.xlsx"
            layout_path = base / "layout.json"
            patterns_path = base / "patterns.json"

            wb = Workbook()
            ws = wb.active
            ws["A2"] = ""
            ws["B2"] = "MG"
            wb.save(workbook_path)
            wb.close()

            layout_path.write_text(json.dumps(layout_payload), encoding="utf-8")
            patterns_path.write_text(json.dumps(patterns_payload), encoding="utf-8")

            layout = Layout.load(layout_path)
            rules = load_patterns(patterns_path)
            errors = validate_workbook(workbook_path, rules, layout)
            payload = grouped_errors_payload(errors)

        self.assertEqual(payload["count"], 2)
        messages = {group["message"] for group in payload["groups"]}
        self.assertIn("Célula vazia", messages)
        self.assertIn("Valor não pertence ao domínio permitido", messages)

    def test_write_error_cell_highlights_marks_error_cells(self) -> None:
        layout_payload = {
            "initialLine": 2,
            "columns": {
                "A": {"description": "CPF", "type": "text", "required": True, "patterns": "ER1"},
                "B": {"description": "UF", "required": True, "domain": "D1"},
            },
        }
        patterns_payload = {
            "patterns": {"ER1": r"^\d{11}$"},
            "domains": {"D1": ["SP", "RJ"]},
        }

        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            workbook_path = base / "sample.xlsx"
            layout_path = base / "layout.json"
            patterns_path = base / "patterns.json"
            out_xlsx = base / "marked.xlsx"

            wb = Workbook()
            ws = wb.active
            ws["A2"] = ""
            ws["B2"] = "MG"
            wb.save(workbook_path)
            wb.close()

            layout_path.write_text(json.dumps(layout_payload), encoding="utf-8")
            patterns_path.write_text(json.dumps(patterns_payload), encoding="utf-8")

            layout = Layout.load(layout_path)
            rules = load_patterns(patterns_path)
            errors = validate_workbook(workbook_path, rules, layout)
            write_error_cell_highlights(workbook_path, layout, errors, out_xlsx)

            wb_marked = load_workbook(out_xlsx, read_only=False, data_only=False)
            try:
                mws = wb_marked.active
                self.assertEqual(mws["A2"].fill.fill_type, "solid")
                self.assertEqual(mws["A2"].fill.start_color.rgb, "FFFFCCCC")
                self.assertEqual(mws["B2"].fill.fill_type, "solid")
                self.assertEqual(mws["B2"].fill.start_color.rgb, "FFFFCCCC")
            finally:
                wb_marked.close()


if __name__ == "__main__":
    unittest.main()
