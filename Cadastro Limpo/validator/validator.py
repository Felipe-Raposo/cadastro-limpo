from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Set, Tuple, Union

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from validator.layout import Layout
from validator.patterns import CompiledRules


def _pattern_ids(rule_ids: List[str]) -> Optional[str]:
    if not rule_ids:
        return None
    return " | ".join(rule_ids)


def _cell_heading(layout: Layout, col_letter: str) -> Optional[str]:
    d = layout.descriptions.get(col_letter)
    return d if d else None


@dataclass
class CellError:
    row: int
    col_letter: str
    heading: Optional[str]
    pattern_ids: Optional[str]
    domain_id: Optional[str]
    value: Optional[str]
    message: str


def grouped_errors_payload(errors: List[CellError]) -> Dict[str, Any]:
    """Monta JSON com count global e groups por (heading do layout, message)."""
    buckets: Dict[Tuple[Optional[str], str], List[CellError]] = defaultdict(list)
    for e in errors:
        buckets[(e.heading, e.message)].append(e)

    def group_sort_key(key: Tuple[Optional[str], str]) -> Tuple[int, str, str]:
        heading, message = key
        if heading is None:
            return (1, "", message)
        return (0, heading, message)

    sorted_keys = sorted(buckets.keys(), key=group_sort_key)
    groups: List[Dict[str, Any]] = []

    for key in sorted_keys:
        items = buckets[key]
        items_sorted = sorted(
            items, key=lambda e: (e.row, column_index_from_string(e.col_letter))
        )
        heading, message = key
        first = items_sorted[0]
        err_list = [{"cell": f"{e.col_letter}{e.row}", "value": e.value} for e in items_sorted]

        g: Dict[str, Any] = {
            "column": heading,
            "message": message,
            "count": len(items_sorted),
            "errors": err_list,
        }
        if first.pattern_ids:
            g["pattern"] = first.pattern_ids
        if first.domain_id:
            g["domain"] = first.domain_id
        groups.append(g)

    return {"count": len(errors), "groups": groups}


def _cell_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    return str(value)


def _is_blank(text: Optional[str]) -> bool:
    if text is None:
        return True
    return text.strip() == ""


def _presence_column_indices(layout: Layout) -> List[int]:
    """Colunas usadas para decidir se a linha tem dados (não é linha em branco)."""
    letters = sorted(layout.rules.keys(), key=column_index_from_string)
    return [column_index_from_string(letter) for letter in letters]


def _row_has_any_value(ws: Worksheet, row: int, col_indices: List[int]) -> bool:
    for col_idx in col_indices:
        v = ws.cell(row=row, column=col_idx).value
        if v is not None and str(v).strip() != "":
            return True
    return False


def _empty_requires_error(col_letter: str, layout: Layout) -> bool:
    """True se célula vazia (ou só espaços) deve gerar erro de obrigatoriedade."""
    return col_letter in layout.required_columns


def _normalize_value_for_type(text: str, column_type: Optional[str]) -> str:
    if column_type is None:
        return text.strip()
    normalized_type = column_type.strip().lower()
    if normalized_type == "text":
        return text.strip()
    return text.strip()


class WorkbookValidationCancelled(Exception):
    """Lançado quando ``cancel_check`` indica cancelamento durante a validação."""


_ERROR_HIGHLIGHT_FILL = PatternFill(
    start_color="FFFFCCCC",
    end_color="FFFFCCCC",
    fill_type="solid",
)


def _worksheet_for_layout(wb: Any, layout: Layout) -> Worksheet:
    if layout.sheet_name:
        if layout.sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Aba {layout.sheet_name!r} não existe. Abas: {wb.sheetnames}"
            )
        return wb[layout.sheet_name]
    return wb.active


def write_error_cell_highlights(
    workbook_path: Union[Path, str],
    layout: Layout,
    errors: List[CellError],
    output_path: Union[Path, str],
) -> None:
    """Grava uma cópia do workbook pintando de vermelho claro cada célula em ``errors``.

    Usa ``data_only=False`` para preservar fórmulas ao salvar. Células que deixaram
    de falhar na validação não têm o preenchimento removido automaticamente.
    """
    path = Path(workbook_path)
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(path, read_only=False, data_only=False)
    try:
        ws = _worksheet_for_layout(wb, layout)
        seen: Set[Tuple[int, str]] = set()
        for e in errors:
            key = (e.row, e.col_letter)
            if key in seen:
                continue
            seen.add(key)
            col_idx = column_index_from_string(e.col_letter)
            ws.cell(row=e.row, column=col_idx).fill = _ERROR_HIGHLIGHT_FILL
        wb.save(out)
    finally:
        wb.close()


def validate_workbook(
    workbook_path: Union[Path, str],
    rules: CompiledRules,
    layout: Layout,
    *,
    on_row_progress: Optional[Callable[[int, int, int], None]] = None,
    cancel_check: Optional[Callable[[], bool]] = None,
) -> List[CellError]:
    path = Path(workbook_path)
    # read_only=True não preenche max_row de forma confiável; data_only usa valores calculados.
    wb = load_workbook(path, read_only=False, data_only=True)
    try:
        ws = _worksheet_for_layout(wb, layout)

        presence_cols = _presence_column_indices(layout)
        column_indices = {
            col_letter: column_index_from_string(col_letter)
            for col_letter in layout.rules
        }
        missing_regexes_by_col: Dict[str, List[str]] = {}
        available_regexes: Set[str] = set(rules.regexes)
        for col_letter, rule_ids in layout.rules.items():
            missing_regexes_by_col[col_letter] = [
                rid for rid in rule_ids if rid not in available_regexes
            ]
        available_domains: Set[str] = set(rules.domains)
        missing_domains_by_col = {
            col_letter: domain_id
            for col_letter, domain_id in layout.domain_rules.items()
            if domain_id not in available_domains
        }

        max_row = ws.max_row or layout.initial_line
        errors: List[CellError] = []
        total_rows = max(0, max_row - layout.initial_line + 1)

        for row in range(layout.initial_line, max_row + 1):
            if cancel_check and cancel_check():
                raise WorkbookValidationCancelled()
            if on_row_progress is not None and total_rows > 0:
                idx = row - layout.initial_line + 1
                on_row_progress(idx - 1, total_rows, row)
            if not _row_has_any_value(ws, row, presence_cols):
                if on_row_progress is not None and total_rows > 0:
                    done = row - layout.initial_line + 1
                    on_row_progress(done, total_rows, row)
                continue

            for col_letter, rule_ids in layout.rules.items():
                if cancel_check and cancel_check():
                    raise WorkbookValidationCancelled()
                domain_id = layout.domain_rules.get(col_letter)
                pid = _pattern_ids(rule_ids)
                col_idx = column_indices[col_letter]
                raw = ws.cell(row=row, column=col_idx).value
                text = _cell_str(raw)
                display = text if text is not None else ""

                if _is_blank(text):
                    if not _empty_requires_error(col_letter, layout):
                        continue
                    errors.append(
                        CellError(
                            row=row,
                            col_letter=col_letter,
                            heading=_cell_heading(layout, col_letter),
                            pattern_ids=pid,
                            domain_id=domain_id,
                            value=None,
                            message="Célula vazia",
                        )
                    )
                    continue

                has_patterns = bool(rule_ids)
                has_domain = domain_id is not None

                if not has_patterns and not has_domain:
                    continue

                if has_patterns:
                    missing = missing_regexes_by_col[col_letter]
                    if missing:
                        errors.append(
                            CellError(
                                row=row,
                                col_letter=col_letter,
                                heading=_cell_heading(layout, col_letter),
                                pattern_ids=pid,
                                domain_id=domain_id,
                                value=display,
                                message="Regra(s) de regex não definida(s) no arquivo de padrões: "
                                + ", ".join(repr(m) for m in missing),
                            )
                        )
                        continue

                    if not any(rules.regexes[rid].fullmatch(text) for rid in rule_ids):
                        errors.append(
                            CellError(
                                row=row,
                                col_letter=col_letter,
                                heading=_cell_heading(layout, col_letter),
                                pattern_ids=pid,
                                domain_id=domain_id,
                                value=display,
                                message="Valor não corresponde a nenhuma das expressões regulares",
                            )
                        )
                        continue

                if has_domain:
                    if col_letter in missing_domains_by_col:
                        errors.append(
                            CellError(
                                row=row,
                                col_letter=col_letter,
                                heading=_cell_heading(layout, col_letter),
                                pattern_ids=pid,
                                domain_id=domain_id,
                                value=display,
                                message=f"Domínio {domain_id!r} não definido no arquivo de padrões",
                            )
                        )
                        continue
                    column_type = layout.column_types.get(col_letter)
                    normalized = _normalize_value_for_type(text, column_type)
                    if normalized not in rules.domains[domain_id]:
                        errors.append(
                            CellError(
                                row=row,
                                col_letter=col_letter,
                                heading=_cell_heading(layout, col_letter),
                                pattern_ids=pid,
                                domain_id=domain_id,
                                value=display,
                                message="Valor não pertence ao domínio permitido",
                            )
                        )

            if on_row_progress is not None and total_rows > 0:
                done = row - layout.initial_line + 1
                on_row_progress(done, total_rows, row)

        return errors
    finally:
        wb.close()
