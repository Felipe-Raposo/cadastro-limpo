"""Regressão: trim em text_columns deve ocorrer em todas as linhas, não só nas amostradas para progresso."""
from __future__ import annotations

import io
import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from sanitiser.layout import SanitiserLayout
from sanitiser.updater import update_workbook_from_api


def _minimal_layout_json() -> str:
    return json.dumps(
        {
            "initialLine": 2,
            "columns": {"B": {"type": "text"}},
            "sanitiser": {
                "entity": {
                    "cpf_cnpj_column": "A",
                    "cpf": {
                        "api": {
                            "url": "http://127.0.0.1:9/unused/{cpf}",
                            "headers": {},
                        },
                        "response_to_column": {},
                    },
                    "cnpj": {
                        "api": {
                            "url": "http://127.0.0.1:9/unused/{cnpj}",
                            "headers": {},
                        },
                        "response_to_column": {},
                    },
                }
            },
        }
    )


class TestUpdaterTrimAllRows(unittest.TestCase):
    def test_trim_applies_on_rows_not_aligned_with_progress_step(self) -> None:
        """Com >200 linhas, step de progresso é 10; trim deve rodar mesmo quando done % step != 0."""
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            in_xlsx = tmp_path / "in.xlsx"
            out_xlsx = tmp_path / "out.xlsx"
            layout_path = tmp_path / "layout.json"
            layout_path.write_text(_minimal_layout_json(), encoding="utf-8")
            layout = SanitiserLayout.load(layout_path)

            wb = Workbook()
            ws = wb.active
            # 201 linhas de dados => total_rows > 200 => step == 10
            for row in range(2, 203):
                ws.cell(row=row, column=1).value = None
                ws.cell(row=row, column=2).value = "  trimmed  "
            wb.save(in_xlsx)
            wb.close()

            update_workbook_from_api(
                in_xlsx,
                layout,
                out_xlsx,
                use_api_cache=False,
                log_stream=io.StringIO(),
                progress_stream=io.StringIO(),
            )

            out_wb = load_workbook(out_xlsx, read_only=True, data_only=True)
            try:
                ows = out_wb.active
                # Linha 5: done = 5 - 2 + 1 = 4; 4 % 10 != 0 e não é primeira/última
                self.assertEqual(ows.cell(row=5, column=2).value, "trimmed")
                # Linha 11: done = 10 => ainda múltiplo de 10; escolher linha 6, done=5
                self.assertEqual(ows.cell(row=6, column=2).value, "trimmed")
                self.assertEqual(ows.cell(row=202, column=2).value, "trimmed")
            finally:
                out_wb.close()


if __name__ == "__main__":
    unittest.main()
