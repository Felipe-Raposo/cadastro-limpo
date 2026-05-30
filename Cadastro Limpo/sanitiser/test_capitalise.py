"""Testes para `capitalise` no leiaute (entity/address) e aplicação nas células."""
from __future__ import annotations

import json
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook

from sanitiser.layout import SanitiserLayout
from sanitiser.updater import _apply_address_mapping, _apply_entity_mapping


def _minimal_layout_dict(**sanitise_overrides: object) -> dict:
    base = {
        "initialLine": 2,
        "sanitiser": {
            "entity": {
                "cpf_cnpj_column": "C",
                "cpf": {
                    "api": {"url": "https://example.invalid/cpf/{cpf}"},
                    "response_to_column": {"NOME": "A"},
                },
                "cnpj": {
                    "api": {"url": "https://example.invalid/cnpj/{cnpj}"},
                    "response_to_column": {"razao_social": "A"},
                },
            }
        },
    }
    st = base["sanitiser"]
    assert isinstance(st, dict)
    for key, val in sanitise_overrides.items():
        st[key] = val
    return base


class TestLayoutCapitalise(unittest.TestCase):
    def _load(self, payload: dict) -> SanitiserLayout:
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".json", encoding="utf-8", delete=False
        ) as tmp:
            json.dump(payload, tmp, ensure_ascii=False)
            path = Path(tmp.name)
        try:
            return SanitiserLayout.load(path)
        finally:
            path.unlink(missing_ok=True)

    def test_defaults_false_when_omitted(self) -> None:
        layout = self._load(_minimal_layout_dict())
        self.assertFalse(layout.entity_capitalise)
        self.assertIsNone(layout.address)

    def test_entity_capitalise_true(self) -> None:
        payload = _minimal_layout_dict()
        ent = payload["sanitiser"]["entity"]
        assert isinstance(ent, dict)
        ent["capitalise"] = True
        layout = self._load(payload)
        self.assertTrue(layout.entity_capitalise)

    def test_address_capitalise_true(self) -> None:
        payload = _minimal_layout_dict(
            address={
                "cep_column": "P",
                "capitalise": True,
                "api": {"url": "https://viacep.com.br/ws/{cep}/json/"},
                "response_to_column": {"localidade": "O"},
            }
        )
        layout = self._load(payload)
        assert layout.address is not None
        self.assertTrue(layout.address.capitalise)

    def test_capitalise_invalid_type_raises(self) -> None:
        payload = _minimal_layout_dict()
        ent = payload["sanitiser"]["entity"]
        assert isinstance(ent, dict)
        ent["capitalise"] = "yes"
        with self.assertRaises(ValueError) as ctx:
            self._load(payload)
        self.assertIn("capitalise", str(ctx.exception).lower())


class TestUpdaterCapitalise(unittest.TestCase):
    def test_entity_custom_key_only_upper_when_flag(self) -> None:
        wb = Workbook()
        ws = wb.active
        mapping = {"complemento": "B"}
        col_idx = {"complemento": 2}
        payload = {"complemento": "texto misto"}

        _apply_entity_mapping(
            ws,
            1,
            payload,
            mapping,
            col_idx,
            use_cpf_aliases=False,
            entity_capitalise=False,
        )
        self.assertEqual(ws.cell(row=1, column=2).value, "texto misto")

        _apply_entity_mapping(
            ws,
            2,
            payload,
            mapping,
            col_idx,
            use_cpf_aliases=False,
            entity_capitalise=True,
        )
        self.assertEqual(ws.cell(row=2, column=2).value, "TEXTO MISTO")

    def test_entity_nasc_still_parsed_with_capitalise(self) -> None:
        wb = Workbook()
        ws = wb.active
        mapping = {"NASC": "B"}
        col_idx = {"NASC": 2}
        payload = {"NASC": "1990-05-13"}

        _apply_entity_mapping(
            ws,
            1,
            payload,
            mapping,
            col_idx,
            use_cpf_aliases=True,
            entity_capitalise=True,
        )
        self.assertEqual(ws.cell(row=1, column=2).value, date(1990, 5, 13))

    def test_address_upper_after_merge(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=5, value="Antiga, bloco 2")
        mapping = {"logradouro": "E"}
        col_idx = {"logradouro": 5}

        _apply_address_mapping(
            ws,
            1,
            {"logradouro": "rua nova"},
            mapping,
            col_idx,
            address_capitalise=True,
        )
        self.assertEqual(ws.cell(row=1, column=5).value, "RUA NOVA, BLOCO 2")

    def test_address_no_upper_when_flag_false(self) -> None:
        wb = Workbook()
        ws = wb.active
        mapping = {"localidade": "C"}
        col_idx = {"localidade": 3}

        _apply_address_mapping(
            ws,
            1,
            {"localidade": "São Paulo"},
            mapping,
            col_idx,
            address_capitalise=False,
        )
        self.assertEqual(ws.cell(row=1, column=3).value, "São Paulo")


if __name__ == "__main__":
    unittest.main()
