from __future__ import annotations

import errno
import json
import platform
import re
import sys
import time
import urllib.error
import urllib.request
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, Mapping, Optional, Tuple, Union

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string

from sanitiser.api_cache import ApiCache, resolve_cache_db_path, source_id_from
from sanitiser.layout import SanitiserLayout

_DIGITS_ONLY = re.compile(r"\D+")

_CPF_LEN = 11
_CNPJ_LEN = 14
_CEP_LEN = 8

# Aliases para respostas de CPF (cpf-brasil e similares).
_CPF_FIELD_ALIASES: Dict[str, Tuple[str, ...]] = {
    "NOME": ("NOME", "nome"),
    "SEXO": ("SEXO", "sexo"),
    "NASC": ("NASC", "nasc", "data_nascimento", "dataNascimento"),
    "NOME_MAE": ("NOME_MAE", "nome_mae", "MAE", "mae", "nomeMae"),
}

_DEFAULT_TIMEOUT_SEC = 30.0
_DEFAULT_MAX_FETCH_RETRIES = 3

# Destaques na planilha de saída (ARGB, mesmo padrão do validator).
_API_OK_FILL = PatternFill(
    start_color="FFCCFFCC",
    end_color="FFCCFFCC",
    fill_type="solid",
)
_API_MISSING_FILL = PatternFill(
    start_color="FFFFE699",
    end_color="FFFFE699",
    fill_type="solid",
)


def _browser_platform_token() -> str:
    """Segmento (platform) de um User-Agent estilo Chrome, derivado do SO/host atual."""
    sys_name = platform.system()
    machine = platform.machine() or "unknown"
    if sys_name == "Darwin":
        mac_release = (platform.mac_ver()[0] or "").strip()
        if not mac_release:
            # fallback: versão do kernel Darwin (ex.: 24.3.0)
            mac_release = platform.release() or "unknown"
        os_x = mac_release.replace(".", "_")
        # Chrome no macOS mantém "Intel Mac OS X" no token por compatibilidade web.
        return f"Macintosh; Intel Mac OS X {os_x}"
    if sys_name == "Windows":
        ver = (platform.version() or "").strip()
        parts = ver.split(".")
        if len(parts) >= 2 and parts[0].isdigit():
            nt = f"Windows NT {parts[0]}.{parts[1]}"
        else:
            nt = "Windows NT 10.0"
        if machine in ("AMD64", "x86_64"):
            return f"{nt}; Win64; x64"
        if machine in ("ARM64",):
            return f"{nt}; Win64; ARM64"
        return f"{nt}; {machine}"
    if sys_name == "Linux":
        return f"X11; Linux {machine}"
    return f"{sys_name}; {machine}"


def _default_user_agent() -> str:
    return (
        f"Mozilla/5.0 ({_browser_platform_token()}) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
    )


# urllib não envia User-Agent de navegador; muitos serviços/WAF respondem 403 a
# "Python-urllib/3.x" enquanto o mesmo URL no Postman funciona. O layout JSON
# pode sobrescrever qualquer chave (mesclagem: defaults primeiro, depois api.headers).
_DEFAULT_HTTP_HEADERS: Dict[str, str] = {
    "User-Agent": _default_user_agent(),
    "Accept": "application/json, text/plain, */*",
}


def _merge_http_headers(headers: Mapping[str, str]) -> Dict[str, str]:
    merged = dict(_DEFAULT_HTTP_HEADERS)
    merged.update(dict(headers))
    return merged


class WorkbookUpdateCancelled(Exception):
    """Lançado quando ``cancel_check`` indica cancelamento durante a atualização."""
_RETRY_BASE_SEC = 0.5

_HTTP_TRANSIENT_CODES = frozenset({408, 425, 429, 500, 502, 503, 504})
_CACHED_HTTP_ERR_STATUS = re.compile(r"^HTTP (\d{3}):")
_TRANSIENT_ERRNOS = frozenset(
    {
        errno.ETIMEDOUT,
        errno.ECONNRESET,
        errno.ECONNREFUSED,
        errno.EPIPE,
        errno.EHOSTUNREACH,
        errno.ENETUNREACH,
    }
)


def _cached_error_is_http_4xx(err: str, http_status: Optional[int] = None) -> bool:
    if http_status is not None and 400 <= http_status <= 499:
        return True
    m = _CACHED_HTTP_ERR_STATUS.match(err)
    return m is not None and 400 <= int(m.group(1)) <= 499


def _is_transient_network_reason(reason: object) -> bool:
    if isinstance(reason, TimeoutError):
        return True
    if isinstance(reason, OSError) and reason.errno in _TRANSIENT_ERRNOS:
        return True
    text = str(reason).lower()
    if "timed out" in text or "timeout" in text:
        return True
    if "temporarily unavailable" in text:
        return True
    return False


def _cell_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        return str(value)
    return str(value)


def _try_as_date(value: Any) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        s = value.strip()
        if len(s) >= 10 and s[4] == "-" and s[7] == "-":
            try:
                y, m, d = int(s[0:4]), int(s[5:7]), int(s[8:10])
                return date(y, m, d)
            except ValueError:
                return None
    return None


def _values_differ_for_highlight(before: Any, after: Any) -> bool:
    """True se o conteúdo da célula muda de forma relevante (destaque verde)."""
    db = _try_as_date(before)
    da = _try_as_date(after)
    if db is not None and da is not None:
        return db != da
    if before == after:
        return False
    sb = (_cell_str(before) or "").strip()
    sa = (_cell_str(after) or "").strip()
    if sb == sa:
        return False
    if not sb and not sa:
        return False
    return True


def _normalize_digits(text: Optional[str]) -> Optional[str]:
    if text is None:
        return None
    digits = _DIGITS_ONLY.sub("", text.strip())
    return digits or None


def _normalize_cep_eight(digits: Optional[str]) -> Optional[str]:
    """CEP com 8 dígitos; se houver mais dígitos, usa os 8 primeiros."""
    if not digits or not digits.isdigit():
        return None
    if len(digits) < _CEP_LEN:
        return None
    return digits[:_CEP_LEN]


def _extract_payload(parsed: Any) -> Optional[Dict[str, Any]]:
    if not isinstance(parsed, dict):
        return None
    if parsed.get("success") is False:
        return None
    data = parsed.get("data")
    if isinstance(data, dict):
        return data
    return parsed if parsed else None


def _build_payload_lookup(payload: Dict[str, Any]) -> Dict[str, Any]:
    return {str(k).upper(): v for k, v in payload.items()}


def _pick_cpf_field(
    payload: Dict[str, Any],
    logical: str,
    *,
    payload_lookup: Optional[Dict[str, Any]] = None,
) -> Optional[str]:
    names = _CPF_FIELD_ALIASES.get(logical, (logical,))
    for name in names:
        if name not in payload:
            continue
        raw = payload[name]
        if raw is None:
            continue
        if isinstance(raw, str):
            s = raw.strip()
            if s:
                return s
        elif isinstance(raw, (int, float)):
            return str(raw)
    upper_payload = payload_lookup or _build_payload_lookup(payload)
    for name in names:
        u = name.upper()
        if u in upper_payload:
            raw = upper_payload[u]
            if isinstance(raw, str) and raw.strip():
                return raw.strip()
    return None


def _pick_mapped_field(
    payload: Dict[str, Any],
    json_key: str,
    *,
    payload_lookup: Optional[Dict[str, Any]] = None,
) -> Optional[str]:
    """Resolve valor no JSON pela chave do leiaute (exata ou ignorando maiúsculas)."""
    if json_key in payload:
        raw = payload[json_key]
        if isinstance(raw, str) and raw.strip():
            return raw.strip()
        if isinstance(raw, (int, float)) and raw is not None:
            return str(raw)
    lookup = payload_lookup or _build_payload_lookup(payload)
    key_upper = json_key.upper()
    if key_upper in lookup:
        raw = lookup[key_upper]
        if isinstance(raw, str) and raw.strip():
            return raw.strip()
        if isinstance(raw, (int, float)) and raw is not None:
            return str(raw)
    return None


def _nasc_value_for_excel(iso_or_text: str) -> Union[str, date, datetime]:
    s = iso_or_text.strip()
    if len(s) == 10 and s[4] == "-" and s[7] == "-":
        try:
            y, m, d = int(s[0:4]), int(s[5:7]), int(s[8:10])
            return date(y, m, d)
        except ValueError:
            pass
    return s


def _entity_key_uppercase(json_key: str) -> bool:
    k = json_key.upper()
    if k in ("NASC", "SEXO"):
        return False
    if k in ("NOME", "NOME_MAE"):
        return True
    if k in ("RAZAO_SOCIAL", "NOME_FANTASIA"):
        return True
    if "RAZAO" in k or "FANTASIA" in k:
        return True
    return False


def _apply_entity_mapping(
    ws,
    row: int,
    payload: Dict[str, Any],
    mapping: Mapping[str, Optional[str]],
    column_indices: Mapping[str, int],
    *,
    use_cpf_aliases: bool,
    entity_capitalise: bool,
) -> None:
    payload_lookup = _build_payload_lookup(payload)
    for json_key, col_letter in mapping.items():
        if not col_letter:
            continue
        if use_cpf_aliases:
            text = _pick_cpf_field(payload, json_key, payload_lookup=payload_lookup)
        else:
            text = _pick_mapped_field(payload, json_key, payload_lookup=payload_lookup)
        if text is None:
            continue
        col_idx = column_indices[json_key]
        cell = ws.cell(row=row, column=col_idx)
        before = cell.value
        if use_cpf_aliases and json_key.upper() == "NASC":
            nasc_in = text.strip().upper() if entity_capitalise else text
            new_val = _nasc_value_for_excel(nasc_in)
        elif entity_capitalise or _entity_key_uppercase(json_key):
            new_val = text.upper()
        else:
            new_val = text
        cell.value = new_val
        if _values_differ_for_highlight(before, new_val):
            cell.fill = _API_OK_FILL


def _trim_text_cells(ws, row: int, text_col_indices: Iterable[int]) -> None:
    for col_idx in text_col_indices:
        cell = ws.cell(row=row, column=col_idx)
        v = cell.value
        if not isinstance(v, str):
            continue
        stripped = v.strip()
        if stripped == v:
            continue
        cell.value = stripped if stripped else None


def _merge_logradouro_cell(api_text: str, current_value: Any) -> str:
    """Substitui só o trecho antes da primeira vírgula; o restante (complemento) permanece."""
    s = _cell_str(current_value) or ""
    if "," in s:
        _, tail = s.split(",", 1)
        return f"{api_text.strip()},{tail}"
    return api_text.strip()


def _apply_address_mapping(
    ws,
    row: int,
    payload: Dict[str, Any],
    mapping: Mapping[str, Optional[str]],
    column_indices: Mapping[str, int],
    *,
    address_capitalise: bool,
) -> None:
    payload_lookup = _build_payload_lookup(payload)
    for json_key, col_letter in mapping.items():
        if not col_letter:
            continue
        text = _pick_mapped_field(payload, json_key, payload_lookup=payload_lookup)
        if text is None:
            continue
        col_idx = column_indices[json_key]
        cell = ws.cell(row=row, column=col_idx)
        before = cell.value
        if json_key.lower() == "logradouro":
            text = _merge_logradouro_cell(text, cell.value)
        if address_capitalise:
            text = text.upper()
        cell.value = text
        if _values_differ_for_highlight(before, text):
            cell.fill = _API_OK_FILL


def _fetch_json_once(
    url: str, headers: Dict[str, str], timeout: float
) -> Tuple[Optional[Dict[str, Any]], Optional[str], bool, Optional[int]]:
    """Uma tentativa HTTP + parse. Retorna (payload, err, transient, http_status)."""
    req = urllib.request.Request(url, method="GET", headers=_merge_http_headers(headers))
    http_ok: Optional[int] = None
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            http_ok = resp.getcode()
            body = resp.read()
            charset = resp.headers.get_content_charset() or "utf-8"
    except TimeoutError as e:
        return None, str(e), True, None
    except urllib.error.HTTPError as e:
        transient = e.code in _HTTP_TRANSIENT_CODES
        try:
            err_body = e.read().decode(e.headers.get_content_charset() or "utf-8", errors="replace")
        except Exception:
            err_body = ""
        return None, f"HTTP {e.code}: {err_body[:500]}", transient, e.code
    except urllib.error.URLError as e:
        reason = e.reason if hasattr(e, "reason") else e
        transient = _is_transient_network_reason(reason)
        msg = str(reason if reason is not None else e)
        return None, msg, transient, None
    except OSError as e:
        transient = _is_transient_network_reason(e)
        return None, str(e), transient, None

    try:
        parsed: Any = json.loads(body.decode(charset))
    except json.JSONDecodeError as e:
        return None, f"Resposta não é JSON válido: {e}", False, http_ok

    if isinstance(parsed, dict) and parsed.get("success") is False:
        msg = parsed.get("message") or parsed.get("error") or "success=false"
        return None, str(msg), False, http_ok

    if isinstance(parsed, dict) and parsed.get("erro"):
        msg = parsed.get("message") or "CEP inválido"
        return None, str(msg), False, http_ok

    payload = _extract_payload(parsed)
    if not payload:
        return None, "Resposta sem objeto de dados utilizável", False, http_ok
    return payload, None, False, http_ok


def _fetch_json(
    url: str,
    headers: Dict[str, str],
    timeout: float,
    *,
    max_retries: int = _DEFAULT_MAX_FETCH_RETRIES,
) -> Tuple[Optional[Dict[str, Any]], Optional[str], bool, Optional[int]]:
    """
    HTTP com retentativas e backoff.
    Retorna (payload, err, cache_error, http_status): se err, cache_error indica se deve gravar em ApiCache.put_error.
    Falhas não-transientes retornam cache_error=True.
    Falhas transitórias esgotadas sem sucesso retornam cache_error=False.
    """
    attempts = max(1, max_retries)
    last_err: Optional[str] = None
    last_http_status: Optional[int] = None
    for attempt in range(attempts):
        payload, err, transient, http_status = _fetch_json_once(url, headers, timeout)
        if err is None:
            return payload, None, False, http_status
        last_err = err
        last_http_status = http_status
        if not transient:
            return None, err, True, http_status
        if attempt < attempts - 1:
            time.sleep(_RETRY_BASE_SEC * (2**attempt))
    return None, last_err, False, last_http_status


def _fetch_json_cached(
    cache: Optional[ApiCache],
    kind: str,
    query_key: str,
    source_id: str,
    url: str,
    headers: Dict[str, str],
    timeout: float,
    *,
    max_retries: int = _DEFAULT_MAX_FETCH_RETRIES,
) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    if cache is not None:
        hit = cache.get(kind, query_key, source_id)
        if hit is not None:
            hit_payload, hit_err, hit_http = hit
            if hit_err is not None and _cached_error_is_http_4xx(hit_err, hit_http):
                cache.delete(kind, query_key, source_id)
            else:
                return hit_payload, hit_err
    payload, err, cache_error, http_status = _fetch_json(
        url, headers, timeout, max_retries=max_retries
    )
    if cache is not None:
        if err is not None and cache_error:
            cache.put_error(
                kind, query_key, source_id, err, http_status=http_status
            )
        elif payload is not None:
            cache.put_success(
                kind, query_key, source_id, payload, http_status=http_status
            )
    return payload, err


def _progress_step(total: int) -> int:
    if total <= 200:
        return 1
    if total <= 2000:
        return 10
    return max(50, total // 200)


def _emit_progress(done: int, total: int, stream, *, inline: bool) -> None:
    pct = (done * 100 // total) if total else 0
    msg = f"Processadas {done} linhas de {total} ({pct}%)."
    if inline:
        # \r: mesmo lugar na linha; \033[K: apaga resto (texto anterior mais longo).
        print(f"\r{msg}\033[K", end="", file=stream, flush=True)
    else:
        print(msg, file=stream, flush=True)


def update_workbook_from_api(
    workbook_path: Union[Path, str],
    layout: SanitiserLayout,
    output_path: Union[Path, str],
    *,
    max_lines: Optional[int] = None,
    timeout_sec: float = _DEFAULT_TIMEOUT_SEC,
    max_fetch_retries: int = _DEFAULT_MAX_FETCH_RETRIES,
    log_stream=sys.stderr,
    progress_stream=sys.stdout,
    on_row_progress: Optional[Callable[[int, int, int], None]] = None,
    cancel_check: Optional[Callable[[], bool]] = None,
    use_api_cache: bool = True,
    cache_db_path: Optional[Path] = None,
) -> int:
    if max_lines is not None and max_lines < 1:
        raise ValueError("max_lines (--lines) deve ser inteiro >= 1")
    if max_fetch_retries < 1:
        raise ValueError("max_fetch_retries deve ser inteiro >= 1")

    path = Path(workbook_path)
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(path, read_only=False, data_only=False)
    cache: Optional[ApiCache] = None
    if use_api_cache:
        db_path = cache_db_path if cache_db_path is not None else resolve_cache_db_path()
        cache = ApiCache(db_path)
    warnings = 0
    inline_progress = False
    should_commit_cache_batch = False
    try:
        if layout.sheet_name:
            if layout.sheet_name not in wb.sheetnames:
                raise ValueError(
                    f"Aba {layout.sheet_name!r} não existe. Abas: {wb.sheetnames}"
                )
            ws = wb[layout.sheet_name]
        else:
            ws = wb.active

        doc_idx = column_index_from_string(layout.cpf_cnpj_column)
        cep_idx: Optional[int] = None
        if layout.address is not None:
            cep_idx = column_index_from_string(layout.address.cep_column)

        cpf_headers = dict(layout.cpf_headers)
        cnpj_headers = dict(layout.cnpj_headers)
        cep_headers: Dict[str, str] = {}
        if layout.address is not None:
            cep_headers = dict(layout.address.headers)

        cpf_col_indices = {
            json_key: column_index_from_string(col_letter)
            for json_key, col_letter in layout.cpf_response_to_column.items()
            if col_letter
        }
        cnpj_col_indices = {
            json_key: column_index_from_string(col_letter)
            for json_key, col_letter in layout.cnpj_response_to_column.items()
            if col_letter
        }
        cep_col_indices: Dict[str, int] = {}
        if layout.address is not None:
            cep_col_indices = {
                json_key: column_index_from_string(col_letter)
                for json_key, col_letter in layout.address.response_to_column.items()
                if col_letter
            }

        max_row = ws.max_row or layout.initial_line
        last_row = max_row
        if max_lines is not None:
            last_row = min(max_row, layout.initial_line + max_lines - 1)

        total_rows = max(0, last_row - layout.initial_line + 1)
        step = _progress_step(total_rows) if total_rows else 1
        isatty = getattr(progress_stream, "isatty", lambda: False)
        inline_progress = bool(isatty()) and total_rows > 0

        cpf_source_id = source_id_from(layout.cpf_url, cpf_headers)
        cnpj_source_id = source_id_from(layout.cnpj_url, cnpj_headers)
        cep_source_id: Optional[str] = None
        if layout.address is not None:
            cep_source_id = source_id_from(
                layout.address.url_template, cep_headers
            )

        text_col_indices: Tuple[int, ...] = tuple(
            column_index_from_string(c) for c in sorted(layout.text_columns)
        )
        if cache is not None:
            cache.begin_batch()

        for row in range(layout.initial_line, last_row + 1):
            if cancel_check and cancel_check():
                raise WorkbookUpdateCancelled()
            done = row - layout.initial_line + 1
            if on_row_progress is not None and total_rows > 0:
                on_row_progress(done - 1, total_rows, row)
            _trim_text_cells(ws, row, text_col_indices)
            if step == 1 or done == 1 or done == total_rows or done % step == 0:
                _emit_progress(done, total_rows, progress_stream, inline=inline_progress)

            raw_doc = ws.cell(row=row, column=doc_idx).value
            doc_digits = _normalize_digits(_cell_str(raw_doc))

            if doc_digits is not None and len(doc_digits) == _CPF_LEN and doc_digits.isdigit():
                url = layout.cpf_url.replace("{cpf}", doc_digits)
                payload, err = _fetch_json_cached(
                    cache,
                    "cpf",
                    doc_digits,
                    cpf_source_id,
                    url,
                    cpf_headers,
                    timeout_sec,
                    max_retries=max_fetch_retries,
                )
                if err is not None:
                    print(f"sanitiser: linha {row} (CPF): {err}", file=log_stream)
                    warnings += 1
                    ws.cell(row=row, column=doc_idx).fill = _API_MISSING_FILL
                elif payload:
                    _apply_entity_mapping(
                        ws,
                        row,
                        payload,
                        layout.cpf_response_to_column,
                        cpf_col_indices,
                        use_cpf_aliases=True,
                        entity_capitalise=layout.entity_capitalise,
                    )

            elif doc_digits is not None and len(doc_digits) == _CNPJ_LEN and doc_digits.isdigit():
                url = layout.cnpj_url.replace("{cnpj}", doc_digits)
                payload, err = _fetch_json_cached(
                    cache,
                    "cnpj",
                    doc_digits,
                    cnpj_source_id,
                    url,
                    cnpj_headers,
                    timeout_sec,
                    max_retries=max_fetch_retries,
                )
                if err is not None:
                    print(f"sanitiser: linha {row} (CNPJ): {err}", file=log_stream)
                    warnings += 1
                    ws.cell(row=row, column=doc_idx).fill = _API_MISSING_FILL
                elif payload:
                    _apply_entity_mapping(
                        ws,
                        row,
                        payload,
                        layout.cnpj_response_to_column,
                        cnpj_col_indices,
                        use_cpf_aliases=False,
                        entity_capitalise=layout.entity_capitalise,
                    )

            if layout.address is not None and cep_idx is not None:
                raw_cep = ws.cell(row=row, column=cep_idx).value
                cep_digits = _normalize_digits(_cell_str(raw_cep))
                cep8 = _normalize_cep_eight(cep_digits)
                if cep8:
                    url = layout.address.url_template.replace("{cep}", cep8)
                    payload, err = _fetch_json_cached(
                        cache,
                        "cep",
                        cep8,
                        cep_source_id,
                        url,
                        cep_headers,
                        timeout_sec,
                        max_retries=max_fetch_retries,
                    )
                    if err is not None:
                        print(f"sanitiser: linha {row} (CEP): {err}", file=log_stream)
                        warnings += 1
                        ws.cell(row=row, column=cep_idx).fill = _API_MISSING_FILL
                    elif payload:
                        _apply_address_mapping(
                            ws,
                            row,
                            payload,
                            layout.address.response_to_column,
                            cep_col_indices,
                            address_capitalise=layout.address.capitalise,
                        )

            if on_row_progress is not None and total_rows > 0:
                on_row_progress(done, total_rows, row)

        wb.save(out)
        should_commit_cache_batch = True
    finally:
        if inline_progress:
            print(file=progress_stream)
        wb.close()
        if cache is not None:
            cache.end_batch(commit=should_commit_cache_batch)
            cache.close()

    return warnings
