"""Destaques de célula (verde/amarelo) após consultas às APIs no update_workbook_from_api."""
from __future__ import annotations

import io
import json
import tempfile
import unittest
from pathlib import Path
from typing import Any, Dict, Optional, Tuple
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from sanitiser.layout import SanitiserLayout
from sanitiser import updater as updater_mod
from sanitiser.updater import update_workbook_from_api


def _fill_rgb(cell: Any) -> Optional[str]:
    fill = cell.fill
    if fill.fill_type != "solid":
        return None
    rgb = fill.start_color.rgb
    if rgb is None:
        return None
    if isinstance(rgb, str):
        return rgb
    return str(rgb)


def _layout_json_with_address(*, include_address: bool) -> str:
    st: Dict[str, Any] = {
        "entity": {
            "cpf_cnpj_column": "A",
            "cpf": {
                "api": {"url": "https://example.invalid/cpf/{cpf}"},
                "response_to_column": {"NOME": "B"},
            },
            "cnpj": {
                "api": {"url": "https://example.invalid/cnpj/{cnpj}"},
                "response_to_column": {"razao_social": "B"},
            },
        }
    }
    if include_address:
        st["address"] = {
            "cep_column": "C",
            "api": {"url": "https://example.invalid/cep/{cep}/"},
            "response_to_column": {"localidade": "D"},
        }
    return json.dumps({"initialLine": 2, "sanitiser": st}, ensure_ascii=False)


class TestSanitiseHighlights(unittest.TestCase):
    def test_green_when_api_changes_mapped_cell(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            layout_path = tmp_path / "layout.json"
            layout_path.write_text(_layout_json_with_address(include_address=False), encoding="utf-8")
            layout = SanitiserLayout.load(layout_path)
            in_xlsx = tmp_path / "in.xlsx"
            out_xlsx = tmp_path / "out.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.cell(row=2, column=1).value = "52998224725"
            ws.cell(row=2, column=2).value = None
            wb.save(in_xlsx)
            wb.close()

            def fake_fetch(
                *_a: Any,
                **_k: Any,
            ) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
                return ({"NOME": "FULANO DA SILVA"}, None)

            with patch.object(updater_mod, "_fetch_json_cached", side_effect=fake_fetch):
                update_workbook_from_api(
                    in_xlsx,
                    layout,
                    out_xlsx,
                    use_api_cache=False,
                    log_stream=io.StringIO(),
                    progress_stream=io.StringIO(),
                )

            out = load_workbook(out_xlsx, read_only=False, data_only=True)
            try:
                ows = out.active
                self.assertEqual(ows.cell(row=2, column=2).value, "FULANO DA SILVA")
                self.assertEqual(_fill_rgb(ows.cell(row=2, column=2)), "FFCCFFCC")
            finally:
                out.close()

    def test_no_green_when_value_unchanged(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            layout_path = tmp_path / "layout.json"
            layout_path.write_text(_layout_json_with_address(include_address=False), encoding="utf-8")
            layout = SanitiserLayout.load(layout_path)
            in_xlsx = tmp_path / "in.xlsx"
            out_xlsx = tmp_path / "out.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.cell(row=2, column=1).value = "52998224725"
            ws.cell(row=2, column=2).value = "FULANO DA SILVA"
            wb.save(in_xlsx)
            wb.close()

            def fake_fetch(
                *_a: Any,
                **_k: Any,
            ) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
                return ({"NOME": "FULANO DA SILVA"}, None)

            with patch.object(updater_mod, "_fetch_json_cached", side_effect=fake_fetch):
                update_workbook_from_api(
                    in_xlsx,
                    layout,
                    out_xlsx,
                    use_api_cache=False,
                    log_stream=io.StringIO(),
                    progress_stream=io.StringIO(),
                )

            out = load_workbook(out_xlsx, read_only=False, data_only=True)
            try:
                ows = out.active
                self.assertEqual(ows.cell(row=2, column=2).value, "FULANO DA SILVA")
                self.assertNotEqual(_fill_rgb(ows.cell(row=2, column=2)), "FFCCFFCC")
            finally:
                out.close()

    def test_yellow_on_cpf_key_cell_when_api_error(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            layout_path = tmp_path / "layout.json"
            layout_path.write_text(_layout_json_with_address(include_address=False), encoding="utf-8")
            layout = SanitiserLayout.load(layout_path)
            in_xlsx = tmp_path / "in.xlsx"
            out_xlsx = tmp_path / "out.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.cell(row=2, column=1).value = "52998224725"
            wb.save(in_xlsx)
            wb.close()

            def fake_fetch(
                *_a: Any,
                **_k: Any,
            ) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
                return (None, "HTTP 404: not found")

            with patch.object(updater_mod, "_fetch_json_cached", side_effect=fake_fetch):
                update_workbook_from_api(
                    in_xlsx,
                    layout,
                    out_xlsx,
                    use_api_cache=False,
                    log_stream=io.StringIO(),
                    progress_stream=io.StringIO(),
                )

            out = load_workbook(out_xlsx, read_only=False, data_only=True)
            try:
                ows = out.active
                self.assertEqual(_fill_rgb(ows.cell(row=2, column=1)), "FFFFE699")
            finally:
                out.close()

    def test_yellow_on_cnpj_key_cell_when_api_error(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            layout_path = tmp_path / "layout.json"
            layout_path.write_text(_layout_json_with_address(include_address=False), encoding="utf-8")
            layout = SanitiserLayout.load(layout_path)
            in_xlsx = tmp_path / "in.xlsx"
            out_xlsx = tmp_path / "out.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.cell(row=2, column=1).value = "11222333000181"
            wb.save(in_xlsx)
            wb.close()

            def fake_fetch(
                *_a: Any,
                **_k: Any,
            ) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
                return (None, "HTTP 404")

            with patch.object(updater_mod, "_fetch_json_cached", side_effect=fake_fetch):
                update_workbook_from_api(
                    in_xlsx,
                    layout,
                    out_xlsx,
                    use_api_cache=False,
                    log_stream=io.StringIO(),
                    progress_stream=io.StringIO(),
                )

            out = load_workbook(out_xlsx, read_only=False, data_only=True)
            try:
                ows = out.active
                self.assertEqual(_fill_rgb(ows.cell(row=2, column=1)), "FFFFE699")
            finally:
                out.close()

    def test_yellow_on_cep_cell_when_api_error(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            layout_path = tmp_path / "layout.json"
            layout_path.write_text(_layout_json_with_address(include_address=True), encoding="utf-8")
            layout = SanitiserLayout.load(layout_path)
            in_xlsx = tmp_path / "in.xlsx"
            out_xlsx = tmp_path / "out.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.cell(row=2, column=1).value = None
            ws.cell(row=2, column=3).value = "01310100"
            wb.save(in_xlsx)
            wb.close()

            def fake_fetch(
                *_a: Any,
                **_k: Any,
            ) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
                return (None, "CEP inválido")

            with patch.object(updater_mod, "_fetch_json_cached", side_effect=fake_fetch):
                update_workbook_from_api(
                    in_xlsx,
                    layout,
                    out_xlsx,
                    use_api_cache=False,
                    log_stream=io.StringIO(),
                    progress_stream=io.StringIO(),
                )

            out = load_workbook(out_xlsx, read_only=False, data_only=True)
            try:
                ows = out.active
                self.assertEqual(_fill_rgb(ows.cell(row=2, column=3)), "FFFFE699")
            finally:
                out.close()


if __name__ == "__main__":
    unittest.main()
